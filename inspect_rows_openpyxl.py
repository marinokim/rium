import openpyxl
import sys
import os

# Add local pylib if needed, though usually packages are installed in env
# sys.path.append(os.path.join(os.getcwd(), 'pylib'))

file_path = './excel/aron_product_upload_consolidated.xlsx'

try:
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # Rows 138-151 (1-based index in openpyxl matches Excel row numbers)
    # Header is row 1
    
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: i for i, name in enumerate(header_row) if name}
    
    print("Headers:", list(headers.keys()))
    
    start_row = 138
    end_row = 151
    
    print(f"Inspecting rows {start_row} to {end_row}...")
    
    for i, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True)):
        current_row_idx = start_row + i
        # Helper to safely get value
        def get_val(col_names):
            for name in col_names:
                if name in headers:
                    return row[headers[name]]
            return None
            
        model_name = get_val(['ModelName', '모델명', '상품명'])
        image_url = get_val(['ImageURL', '이미지URL'])
        detail_url = get_val(['DetailURL', '상세페이지URL'])
        
        print(f"Row {current_row_idx}: Name='{model_name}', Img='{image_url}', Detail='{detail_url}'")

except Exception as e:
    print(f"Error: {e}")
