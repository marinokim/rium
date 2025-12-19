import sys
import os
sys.path.append(os.path.join(os.getcwd(), 'pylib'))
import openpyxl

file_path = './excel/aron_product_upload_consolidated.xlsx'

try:
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {name: i for i, name in enumerate(header_row) if name}
    
    start_row = 1657
    end_row = 1669
    
    print(f"Inspecting rows {start_row} to {end_row}...")
    
    for i, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True)):
        current_row_idx = start_row + i
        def get_val(col_names):
            for name in col_names:
                if name in headers:
                    return row[headers[name]]
            return None
            
        model_name = get_val(['ModelName', '모델명', '상품명'])
        category = get_val(['Category', '카테고리'])
        brand = get_val(['Brand', '브랜드', '제조사'])
        
        print(f"Row {current_row_idx}: Name='{model_name}', Category='{category}', Brand='{brand}'")

except Exception as e:
    print(f"Error: {e}")
