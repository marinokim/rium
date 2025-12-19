import openpyxl
import pg8000.native
import ssl
import re

def normalize_name(name):
    if not name:
        return ""
    return str(name).replace('_x000D_', '').replace('\r', '').replace('\n', '').strip().lower()

def clean_url(val):
    if not val:
        return ""
    s = str(val)
    # If it contains <img src="...">, extract the src
    # Regex: src=["']([^"']+)["']
    match = re.search(r'src=["\']([^"\']+)["\']', s)
    if match:
        return match.group(1)
    # If it starts with < but no src found, return empty to be safe?
    # Or return as is if it's just a URL
    if s.strip().startswith('<') and not match:
        return "" 
    return s.strip()

def export_to_excel():
    # 1. Load Excel
    file_path = './excel/aron_product_upload_consolidated.xlsx'
    print(f"Loading Excel: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # 2. Map Headers
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[col_name] = idx # 0-indexed column index
            
    # Find target columns
    # ImageURL: 'ImageURL', '이미지URL'
    # DetailURL: 'DetailURL', '상세페이지URL'
    img_col_idx = headers.get('ImageURL') or headers.get('이미지URL')
    dtl_col_idx = headers.get('DetailURL') or headers.get('상세페이지URL')
    name_col_idx = headers.get('ModelName') or headers.get('모델명')
    brand_col_idx = headers.get('Brand') or headers.get('브랜드')
    
    if img_col_idx is None or dtl_col_idx is None or name_col_idx is None:
        print("Error: Could not find required columns (ImageURL, DetailURL, ModelName)")
        return

    # 3. Fetch Data from DB
    print("Fetching products from DB...")
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    con = pg8000.native.Connection(
        user="postgres.zgqnhbaztgpekerhxwbc",
        password="gmlrhks0528&*",
        host="aws-1-ap-northeast-2.pooler.supabase.com",
        port=5432,
        database="postgres",
        ssl_context=ssl_context
    )
    
    # Fetch all relevant products (or filters?)
    # Fetch ID, ModelName, ImageURL, DetailURL
    rows = con.run("SELECT id, model_name, image_url, detail_url FROM products")
    
    # Create lookup map: NormalizedName -> (ImageURL, DetailURL)
    db_map = {}
    for r in rows:
        name = normalize_name(r[1])
        if name:
            db_map[name] = (r[2], r[3])
            
    con.close()
    print(f"Loaded {len(db_map)} products from DB.")

    # 4. Update Excel
    updated_count = 0
    # Iterate from row 2
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        # row is a tuple of cells
        name_cell = row[name_col_idx]
        name_val = normalize_name(name_cell.value)
        
        if name_val in db_map:
            db_img, db_dtl = db_map[name_val]
            
            # Clean URLs
            clean_img = clean_url(db_img)
            clean_dtl = clean_url(db_dtl)
            
            # Update cells if DB has value
            # Only update if DB has non-empty value? 
            # User wants backup, so if DB has value, put it in Excel.
            
            if clean_img:
                # +1 because openpyxl cell is 1-indexed for column? 
                # row[i] gives the cell. We can just set value.
                sheet.cell(row=row_idx, column=img_col_idx+1).value = clean_img
                
            if clean_dtl:
                sheet.cell(row=row_idx, column=dtl_col_idx+1).value = clean_dtl
                
            updated_count += 1
            
    print(f"Updated {updated_count} rows in Excel.")
    
    # 5. Save
    wb.save(file_path)
    print("Excel saved.")

if __name__ == "__main__":
    export_to_excel()
