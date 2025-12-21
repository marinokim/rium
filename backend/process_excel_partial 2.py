import sys
import os
import ssl
sys.path.append(os.path.join(os.getcwd(), 'pylib'))

import openpyxl
import pg8000.native

# Helper to sanitize and parse
def sanitize(val):
    if val is None:
        return None
    s = str(val).replace('_x000D_', '').replace('\r', '').replace('"', '').strip()
    # Reject HTML tags for URLs or Names
    if s.startswith('<') or 'img src' in s:
        return None
    return s

def parse_price(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').replace('원', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

# Helper to normalize names for matching
def normalize_name(name):
    if not name:
        return ""
    # Remove _x000D_, carriage returns, newlines, and extra spaces
    return str(name).replace('_x000D_', '').replace('\r', '').replace('\n', '').strip().lower()

def process_partial_upload():
    # DB Connection
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE

    print("Connecting to DB...")
    con = pg8000.native.Connection(
        user="postgres.zgqnhbaztgpekerhxwbc",
        password="gmlrhks0528&*",
        host="aws-1-ap-northeast-2.pooler.supabase.com",
        port=5432,
        database="postgres",
        ssl_context=ssl_context
    )

    # 1. Pre-fetch all available products to build a robust map
    print("Pre-fetching existing products...")
    all_products = con.run("SELECT id, model_no, model_name FROM products")
    
    # Map: normalized_name -> list of IDs (sorted by ID to ensure deterministic order)
    name_map = {}
    model_no_map = {}
    
    for pid, mno, mname in all_products:
        # Map by Model No
        if mno:
            clean_mno = str(mno).strip()
            model_no_map[clean_mno] = pid
            
        # Map by Name (Sequential)
        if mname:
            n_name = normalize_name(mname)
            if n_name not in name_map:
                name_map[n_name] = []
            name_map[n_name].append(pid)
    
    # Sort ID lists for consistent consumption
    for n in name_map:
        name_map[n].sort()

    print(f"Loaded {len(all_products)} products into map.")

    # Load Excel
    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Get headers
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[str(col_name).strip()] = idx

    print(f"Headers found: {list(headers.keys())}")

    # Process rows 115 to 137
    row_idx = 114
    updated_count = 0
    
    for row in sheet.iter_rows(min_row=115, max_row=137, values_only=True):
        row_idx += 1
        
        # Helper to get value by header name(s)
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    return row[headers[h]]
            return None

        # Truncate helper
        def truncate(val, length=255):
            if val and len(val) > length:
                print(f"Warning: Truncating value '{val[:20]}...' to {length} chars")
                return val[:length]
            return val

        brand = sanitize(get_val(['Brand', '브랜드']))
        model_name_raw = get_val(['ModelName', '모델명'])
        model_name = truncate(sanitize(model_name_raw)) # Truncate Model Name
        model_name_norm = normalize_name(model_name_raw)
        
        model_no = truncate(sanitize(get_val(['ModelNo', '모델번호']))) # Truncate Model No
        
        # Prices
        b2b_price = parse_price(get_val(['B2BPrice', '실판매가', '판매가', 'B2B가']))
        consumer_price = parse_price(get_val(['ConsumerPrice', '소비자가', '소가']))
        supply_price = parse_price(get_val(['SupplyPrice', '공급가', '매입가']))
        if supply_price == 0 and b2b_price > 0:
            supply_price = b2b_price
            
        # Carton Quantity
        quantity_per_carton = parse_price(get_val(['QuantityPerCarton', '카톤수량']))
        if quantity_per_carton == 0:
            quantity_per_carton = 1
            
        # Other fields
        description = sanitize(get_val(['Description', '상세설명']))
        if description: 
             description = description.replace('_x000D_', '\n') # Clean description too
             
        detail_url = truncate(sanitize(get_val(['DetailURL', '상세페이지URL'])))
        image_url = truncate(sanitize(get_val(['ImageURL', '이미지URL'])))

        if not model_name: 
            print(f"Row {row_idx}: Skipping, no model name")
            continue

        # FIND TARGET ID
        target_id = None
        
        # 1. Try Model No
        if model_no and model_no in model_no_map:
            target_id = model_no_map[model_no]
            # (Don't consume from name map? Or should we? Let's assume model_no is authoritative)
        
        # 2. Try Name (Consume from map)
        if not target_id and model_name_norm in name_map:
            ids = name_map[model_name_norm]
            if ids:
                target_id = ids.pop(0) # Consume the first one
                # If empty afterwards, remove? No need.
        
        try:
            if target_id:
                # Update
                # Force Category to 'Gift Set' (ID 8) for this batch as requested
                cat_id = 8 
                
                con.run(
                    """UPDATE products 
                       SET model_name = :m_name,
                           model_no = :m_no,
                           category_id = :cat_id,
                           b2b_price = :b2b, 
                           consumer_price = :consumer, 
                           supply_price = :supply,
                           quantity_per_carton = :qty,
                           description = COALESCE(:desc, description),
                           detail_url = COALESCE(:detail, detail_url),
                           image_url = COALESCE(:img, image_url),
                           updated_at = CURRENT_TIMESTAMP
                       WHERE id = :id""",
                    m_name=model_name,
                    m_no=model_no,
                    cat_id=cat_id,
                    b2b=b2b_price, consumer=consumer_price, supply=supply_price, 
                    qty=quantity_per_carton, desc=description, detail=detail_url, 
                    img=image_url, id=target_id
                )
                print(f"Row {row_idx}: Matched ID {target_id} - Updated Category/ModelNo/Name/Prices")
                updated_count += 1
            else:
                # Insert new product
                # 1. Force Category to 'Gift Set' (ID 8) as requested
                # logic: if it's the specific batch or just overwrite for now.
                # User asked to redo 115-137 as Gift Set.
                cat_id = 8 # Gift Set
                
                # 2. Prepare other fields
                stock_quantity = parse_price(get_val(['Stock', '재고']))
                shipping_fee = parse_price(get_val(['ShippingFee', '배송비']))
                shipping_fee_ind = parse_price(get_val(['ShippingFeeIndividual', '개별배송비']))
                if shipping_fee_ind == 0 and shipping_fee > 0: shipping_fee_ind = shipping_fee
                shipping_fee_carton = parse_price(get_val(['ShippingFeeCarton', '카톤배송비']))
                manufacturer = sanitize(get_val(['Manufacturer', '제조사']))
                origin = sanitize(get_val(['Origin', '원산지']))
                product_spec = sanitize(get_val(['ProductSpec', '제품규격']))
                product_options = sanitize(get_val(['ProductOptions', '옵션']))
                remarks = sanitize(get_val(['Remark', 'remark', '비고']))
                is_tax_free = (str(get_val(['IsTaxFree', '면세여부'])) in ['TRUE', '면세'])

                con.run(
                    """INSERT INTO products (
                        category_id, brand, model_name, description, image_url, b2b_price, stock_quantity,
                        detail_url, consumer_price, supply_price, quantity_per_carton, shipping_fee,
                        manufacturer, origin, is_tax_free, shipping_fee_individual, shipping_fee_carton,
                        product_options, model_no, product_spec, remarks, is_available
                    ) VALUES (
                        :cat_id, :brand, :model_name, :desc, :img, :b2b, :stock,
                        :detail, :consumer, :supply, :qty, :ship,
                        :manuf, :origin, :tax, :ship_ind, :ship_carton,
                        :opts, :m_no, :spec, :rem, true
                    ) RETURNING id""",
                    cat_id=cat_id, brand=brand, model_name=model_name, desc=description, img=image_url,
                    b2b=b2b_price, stock=stock_quantity, detail=detail_url, consumer=consumer_price,
                    supply=supply_price, qty=quantity_per_carton, ship=shipping_fee, manuf=manufacturer,
                    origin=origin, tax=is_tax_free, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    opts=product_options, m_no=model_no, spec=product_spec, rem=remarks
                )
                print(f"Row {row_idx}: Inserted New Product ({model_name})")
                updated_count += 1

        except Exception as e:
            print(f"Row {row_idx}: Error - {e}")

    print(f"Finished. Updated {updated_count} products.")
    con.close()

if __name__ == "__main__":
    process_partial_upload()
