import sys
import os
sys.path.append(os.path.join(os.getcwd(), 'pylib'))
import openpyxl

def sanitize(val):
    if val is None: return None
    return str(val).strip()

def parse_price(val):
    if val is None: return 0
    s = str(val).replace(',', '').replace('원', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

def debug_1081():
    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(col_name).strip(): idx for idx, col_name in enumerate(header_row) if col_name}
    
    print(f"Headers: {headers}")
    
    for row in sheet.iter_rows(min_row=1081, max_row=1081, values_only=True):
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    print(f"Found header '{h}' at index {headers[h]}. Value: {row[headers[h]]}")
                    return row[headers[h]]
            print(f"Headers {possible_headers} not found")
            return None
            
        raw_ind = get_val(['ShippingFeeIndividual', '개별배송비'])
        val_ind = parse_price(raw_ind)
        
        raw_carton = get_val(['ShippingFeeCarton', '카톤배송비'])
        val_carton = parse_price(raw_carton)
        
        print(f"Parsed Ind: {val_ind} (Raw: {raw_ind})")
        print(f"Parsed Carton: {val_carton} (Raw: {raw_carton})")

if __name__ == "__main__":
    debug_1081()
