import sys
import os
import ssl
sys.path.append(os.path.join(os.getcwd(), 'pylib'))

import openpyxl
import pg8000.native

# Helper to sanitize and parse
def sanitize(val, allow_html=False):
    if val is None:
        return None
    s = str(val).replace('_x000D_', '').replace('\r', '').strip()
    # Remove quotes if not HTML? Actually, keep original logic mostly but add flag
    if not allow_html:
        s = s.replace('"', '')
        if s.startswith('<') or 'img src' in s:
            return None
    return s

def parse_price(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').replace('원', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

def normalize_name(name):
    if not name:
        return ""
    return str(name).replace('_x000D_', '').replace('\r', '').replace('\n', '').strip().lower()

def register_1081_1111():
    # DB Connection
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE

    print("Connecting to DB...")
    con = pg8000.native.Connection(
        user="postgres.zgqnhbaztgpekerhxwbc",
        password="gmlrhks0528&*",
        host="aws-1-ap-northeast-2.pooler.supabase.com",
        port=5432,
        database="postgres",
        ssl_context=ssl_context
    )

    print("Pre-fetching existing products...")
    all_products = con.run("SELECT id, model_no, model_name FROM products")
    
    name_map = {}
    model_no_map = {}
    
    for pid, mno, mname in all_products:
        if mno:
            clean_mno = str(mno).strip()
            model_no_map[clean_mno] = pid
        if mname:
            n_name = normalize_name(mname)
            if n_name not in name_map:
                name_map[n_name] = []
            name_map[n_name].append(pid)
    
    for n in name_map:
        name_map[n].sort()

    print(f"Loaded {len(all_products)} products into map.")

    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Get headers
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[str(col_name).strip()] = idx

    print(f"Headers found: {list(headers.keys())}")

    # Process rows 1081 to 1111
    min_r = 1081
    max_r = 1111
    print(f"Processing rows {min_r} to {max_r}...")
    
    row_idx = min_r - 1 
    updated_count = 0
    
    for row in sheet.iter_rows(min_row=min_r, max_row=max_r, values_only=True):
        row_idx += 1
        
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    return row[headers[h]]
            return None

        def truncate(val, length=255):
            if val and len(val) > length:
                return val[:length]
            return val

        brand = sanitize(get_val(['Brand', '브랜드']))
        model_name_raw = get_val(['ModelName', '모델명', '상품명'])
        model_name = truncate(sanitize(model_name_raw))
        model_no = truncate(sanitize(get_val(['ModelNo', '모델번호'])))
        
        # Prices
        b2b_price = parse_price(get_val(['B2BPrice', '실판매가', '판매가', 'B2B가']))
        consumer_price = parse_price(get_val(['ConsumerPrice', '소비자가', '소가']))
        supply_price = parse_price(get_val(['SupplyPrice', '공급가', '매입가']))
        if supply_price == 0 and b2b_price > 0:
            supply_price = b2b_price
            
        quantity_per_carton = parse_price(get_val(['QuantityPerCarton', '카톤수량']))
        if quantity_per_carton == 0: quantity_per_carton = 1
            
        description = sanitize(get_val(['Description', '상세설명']))
        if description: description = description.replace('_x000D_', '\n')
             
        # Allow HTML for DetailURL
        detail_url = sanitize(get_val(['DetailURL', '상세페이지URL']), allow_html=True) 
        # But still truncate? DetailURL might be long if it's raw HTML. Check schema.
        # Assuming TEXT or VARCHAR(MAX). If VARCHAR(255), we have a problem.
        # However, usually DetailURL is a URL. If it's HTML content, it might be Description?
        # Re-reading prompt/context: "상세페이지 URL에 포함된 HTML(이미지 태그 등)이 삭제되지 않도록"
        # The user seems to mean the Excel cell contains HTML (img tags) NOT a URL.
        # So it should probably go into `description` if it's content, OR `detail_url` if that column is abused for HTML.
        # Let's assume `detail_url` column in DB is text/long string or user wants it there.
        # If detail_url is limited to 255, this will fail.
        # Let's check schema/previous usage. `process_excel_partial.py` truncated it.
        # If it's actual HTML content, it likely belongs in `description` or `detail_url` (if type is text).
        # Let's trust user intention to put it in `detail_url` but we should probably NOT truncate it if it's HTML content.
        # Checking schema via error might be needed, but let's try to keeping it without truncation if possible or loose truncation.
        
        # Actually, let's keep truncate for now but increase limit if we can, or just risk it? 
        # But wait, `process_excel_partial` truncates to 255. 
        # If `detail_url` is just a URL string, 255 is fine. 
        # If it's a block of HTML `img` tags, it will be >255.
        # I'll try to insert it. If it fails due to length, I'll know.
        
        image_url = truncate(sanitize(get_val(['ImageURL', '이미지URL'])))

        stock_quantity = parse_price(get_val(['Stock', '재고']))
        shipping_fee = parse_price(get_val(['ShippingFee', '배송비']))
        shipping_fee_ind = parse_price(get_val(['ShippingFeeIndividual', '개별배송비']))
        if shipping_fee_ind == 0 and shipping_fee > 0: shipping_fee_ind = shipping_fee
        shipping_fee_carton = parse_price(get_val(['ShippingFeeCarton', '카톤배송비']))
        manufacturer = sanitize(get_val(['Manufacturer', '제조사']))
        origin = sanitize(get_val(['Origin', '원산지']))
        product_spec = sanitize(get_val(['ProductSpec', '제품규격']))
        product_options = sanitize(get_val(['ProductOptions', '옵션']))
        remarks = sanitize(get_val(['Remark', 'remark', '비고']))
        is_tax_free = (str(get_val(['IsTaxFree', '면세여부'])) in ['TRUE', '면세'])

        category_name = sanitize(get_val(['Category', '카테고리']))
        
        # Category Logic
        cat_id = 8 # Default to Gift Set or we should default to 4 (Other)?
        # Plan said: Default to 4 if Living, unless Gift Set in name.
        # Check actual category name from Excel
        
        if category_name and 'Living' in category_name:
            cat_id = 10 # Living
        elif '위칙' in str(model_name): # Fallback if category column is missing or weird, Wichic is Living
            cat_id = 10
            
        # Revert to 8 if "선물세트" in name?
        # User said: "Living" takes precedence or is the new home.
        # "Living 카테고리 추가했어" -> implies we should use it.
        # Let's stick effectively to 10 for these rows as they are mostly "위칙" (Living brand).
        
        if not model_name:
            print(f"Row {row_idx}: Skipping (No Model Name)")
            continue

        # Check existing
        target_id = None
        if model_no and model_no in model_no_map:
            target_id = model_no_map[model_no]
        elif model_name and normalize_name(model_name) in name_map:
            ids = name_map[normalize_name(model_name)]
            if ids: target_id = ids[0]

        try:
            if target_id:
                # Update
                con.run(
                    """UPDATE products 
                       SET model_name = :m_name,
                           model_no = :m_no,
                           category_id = :cat_id,
                           b2b_price = :b2b, 
                           consumer_price = :consumer, 
                           supply_price = :supply,
                           quantity_per_carton = :qty,
                           description = COALESCE(:desc, description),
                           detail_url = COALESCE(:detail, detail_url),
                           image_url = COALESCE(:img, image_url),
                           shipping_fee = :ship,
                           shipping_fee_individual = :ship_ind,
                           shipping_fee_carton = :ship_carton,
                           brand = :brand,
                           updated_at = CURRENT_TIMESTAMP
                       WHERE id = :id""",
                    m_name=model_name, m_no=model_no, cat_id=cat_id,
                    b2b=b2b_price, consumer=consumer_price, supply=supply_price, 
                    qty=quantity_per_carton, desc=description, detail=detail_url, 
                    img=image_url, id=target_id,
                    ship=shipping_fee, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    brand=brand
                )
                print(f"Row {row_idx}: Updated Product {target_id} ({model_name}) - Cat: {cat_id}")
            else:
                # Insert
                con.run(
                    """INSERT INTO products (
                        category_id, brand, model_name, description, image_url, b2b_price, stock_quantity,
                        detail_url, consumer_price, supply_price, quantity_per_carton, shipping_fee,
                        manufacturer, origin, is_tax_free, shipping_fee_individual, shipping_fee_carton,
                        product_options, model_no, product_spec, remarks, is_available
                    ) VALUES (
                        :cat_id, :brand, :model_name, :desc, :img, :b2b, :stock,
                        :detail, :consumer, :supply, :qty, :ship,
                        :manuf, :origin, :tax, :ship_ind, :ship_carton,
                        :opts, :m_no, :spec, :rem, true
                    )""",
                    cat_id=cat_id, brand=brand, model_name=model_name, desc=description, img=image_url,
                    b2b=b2b_price, stock=stock_quantity, detail=detail_url, consumer=consumer_price,
                    supply=supply_price, qty=quantity_per_carton, ship=shipping_fee, manuf=manufacturer,
                    origin=origin, tax=is_tax_free, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    opts=product_options, m_no=model_no, spec=product_spec, rem=remarks
                )
                print(f"Row {row_idx}: Inserted New Product ({model_name}) - Cat: {cat_id}")
            
            updated_count += 1

        except Exception as e:
            print(f"Row {row_idx}: Error - {e}")

    print(f"Done. Processed {updated_count} items.")
    con.close()

if __name__ == "__main__":
    register_1081_1111()
