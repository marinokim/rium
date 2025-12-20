import sys
import os
import re
import ssl
sys.path.append(os.path.join(os.getcwd(), 'pylib'))

import openpyxl
import pg8000.native

def extract_src(val):
    if not val:
        return None
    s = str(val).strip()
    # If it's a simple URL
    if s.lower().startswith('http') and '<' not in s:
        return s
    # If it's an img tag
    match = re.search(r'src=["\']([^"\']+)["\']', s)
    if match:
        return match.group(1)
    return None

def fix_images():
    # DB Connection
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE

    print("Connecting to DB...")
    con = pg8000.native.Connection(
        user="postgres.zgqnhbaztgpekerhxwbc",
        password="gmlrhks0528&*",
        host="aws-1-ap-northeast-2.pooler.supabase.com",
        port=5432,
        database="postgres",
        ssl_context=ssl_context
    )
    
    # We will iterate the range 2084-2145 where we saw the issue.
    # We can also check 1799-1839 just in case.
    # Let's verify broad range for safety: 1799 - 2145
    min_r = 1799
    max_r = 2145

    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # Get headers to find ImageURL and ModelName/No for matching
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[str(col_name).strip()] = idx
            
    print(f"Processing rows {min_r}-{max_r} to fix image URLs...")
    
    updated_count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=min_r, max_row=max_r, values_only=True), start=min_r):
        
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    return row[headers[h]]
            return None

        raw_img = get_val(['ImageURL', '이미지URL'])
        parsed_url = extract_src(raw_img)
        
        model_name = get_val(['ModelName', '모델명', '상품명'])
        if not model_name: 
            continue
            
        clean_name = str(model_name).replace('_x000D_', '').strip()
        
        if parsed_url:
            # Update DB where image_url is missing
            # We can force update if the new logic gives a valid URL
            query = """
                UPDATE products 
                SET image_url = :img 
                WHERE model_name = :m_name AND (image_url IS NULL OR image_url = '')
            """
            res = con.run(query, img=parsed_url, m_name=clean_name)
            # pg8000 run doesn't return affected count easily in all versions, 
            # but we can try to assume it worked.
            
            # Or we can check if we actually extracted something different
            # For logging:
            if '<img' in str(raw_img):
                 # print(f"Row {row_idx}: Extracted {parsed_url} from tag")
                 pass

    # Verify how many still null
    remain = con.run(f"SELECT count(*) FROM products WHERE (image_url IS NULL OR image_url = '') AND id IN (SELECT id FROM products ORDER BY id DESC LIMIT 100)")
    print(f"Remaining products with null image_url (recent 100): {remain[0][0]}")

    con.close()

if __name__ == "__main__":
    fix_images()
