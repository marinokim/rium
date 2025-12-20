import sys
import os
import ssl
import time
import argparse

# Fix path to pylib (parent dir of backend, or sibling?)
# Struct: root/backend/script.py and root/pylib
# So pylib is at ../pylib relative to this script
current_dir = os.path.dirname(os.path.abspath(__file__))
pylib_path = os.path.join(current_dir, '..', 'pylib')
sys.path.append(pylib_path)

import openpyxl
import pg8000.native
import re

# Helper functions
def sanitize(val):
    if val is None:
        return None
    s = str(val).replace('_x000D_', '').replace('\r', '').replace('"', '').strip()
    return s

def sanitize_html(val):
    if val is None:
        return None
    s = str(val).replace('_x000D_', '\n').strip()
    return s

def parse_price(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').replace('원', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

def normalize_name(name):
    if not name:
        return ""
    return str(name).replace('_x000D_', '').replace('\r', '').replace('\n', '').strip().lower()

def extract_image_url(val):
    if val is None:
        return None
    s = str(val).strip()
    # Check for <img src=...> or <img src="...">
    # Regex explains:
    # <img : match tag start
    # [^>]+ : anything not >
    # src= : match src attribute
    # ["\']? : optional quote
    # ([^"\' >]+) : capture group for URL (stop at quote or space or >)
    match = re.search(r'<img[^>]+src=["\']?([^"\' >]+)["\']?[^>]*>', s, re.IGNORECASE)
    if match:
        return match.group(1).replace('_x000D_', '').strip()
    return sanitize(val)

def register_range(start_row, end_row):
    print(f"Starting registration for range {start_row}-{end_row}")

    # DB Connection
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE

    print("Connecting to DB...")
    con = None
    for i in range(3):
        try:
            con = pg8000.native.Connection(
                user="postgres.zgqnhbaztgpekerhxwbc",
                password="gmlrhks0528&*",
                host="aws-1-ap-northeast-2.pooler.supabase.com",
                port=5432,
                database="postgres",
                ssl_context=ssl_context
            )
            break
        except Exception as e:
            print(f"Connection failed ({i+1}/3): {e}")
            time.sleep(2)
            
    if not con:
        print("Failed to connect to DB.")
        sys.exit(1)

    # Pre-fetch existing products
    print("Pre-fetching existing products...")
    all_products = con.run("SELECT id, model_no, model_name FROM products")
    
    name_map = {}
    model_no_map = {}
    
    for pid, mno, mname in all_products:
        if mno:
            clean_mno = str(mno).strip()
            model_no_map[clean_mno] = pid
        if mname:
            n_name = normalize_name(mname)
            if n_name not in name_map:
                name_map[n_name] = []
            name_map[n_name].append(pid)

    # Excel Load
    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    if not os.path.exists(excel_path):
        # Fallback for when running from backend dir or root
        if os.path.exists('../excel/aron_product_upload_consolidated.xlsx'):
            excel_path = '../excel/aron_product_upload_consolidated.xlsx' 
    
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[str(col_name).strip()] = idx

    print(f"Processing rows {start_row} to {end_row}...")
    updated_count = 0
    inserted_count = 0
    error_count = 0

    # Cache for categories to avoid repeated DB hits
    category_cache = {}

    def get_category_id(cat_name):
        if not cat_name:
            return None
        
        # Normalize: 'Living' -> 'Living'
        cat_key = cat_name.strip().lower()
        if cat_key in category_cache:
            return category_cache[cat_key]

        # Query DB
        rows = con.run("SELECT id FROM categories WHERE lower(name) = :name OR lower(slug) = :slug", name=cat_key, slug=cat_key)
        if rows:
            cat_id = rows[0][0]
            category_cache[cat_key] = cat_id
            return cat_id
        
        # Fallback mapping or creation logic could go here, but for now strict
        # Try some common mappings if needed, or just return None (will error or use default)
        return None


    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    return row[headers[h]]
            return None
        
        def truncate(val, length=255):
            if val and len(val) > length:
                return val[:length]
            return val

        model_name_raw = get_val(['ModelName', '모델명', '상품명'])
        if not model_name_raw:
            continue

        model_name = truncate(sanitize(model_name_raw))
        
        try:
            brand = sanitize(get_val(['Brand', '브랜드', '제조사']))
            model_no = truncate(sanitize(get_val(['ModelNo', '모델번호'])))
            
            # Category
            category_name = sanitize(get_val(['Category', '카테고리']))
            cat_id = get_category_id(category_name)
            
            # Prices
            b2b_price = parse_price(get_val(['B2BPrice', '실판매가', '판매가', 'B2B가']))
            consumer_price = parse_price(get_val(['ConsumerPrice', '소비자가', '소가']))
            supply_price = parse_price(get_val(['SupplyPrice', '공급가', '매입가']))
            if supply_price == 0 and b2b_price > 0:
                supply_price = b2b_price
                
            quantity_per_carton = parse_price(get_val(['QuantityPerCarton', '카톤수량']))
            if quantity_per_carton == 0: quantity_per_carton = 1
                
            description = sanitize_html(get_val(['Description', '상세설명']))
            detail_url = sanitize_html(get_val(['DetailURL', '상세페이지URL']))
            image_url = truncate(extract_image_url(get_val(['ImageURL', '이미지URL'])))

            stock_quantity = parse_price(get_val(['Stock', '재고']))
            shipping_fee = parse_price(get_val(['ShippingFee', '배송비']))
            shipping_fee_ind = parse_price(get_val(['ShippingFeeIndividual', '개별배송비']))
            if shipping_fee_ind == 0 and shipping_fee > 0: shipping_fee_ind = shipping_fee
            shipping_fee_carton = parse_price(get_val(['ShippingFeeCarton', '카톤배송비']))
            manufacturer = sanitize(get_val(['Manufacturer', '제조사']))
            origin = sanitize(get_val(['Origin', '원산지']))
            product_spec = sanitize(get_val(['ProductSpec', '제품규격']))
            product_options = sanitize(get_val(['ProductOptions', '옵션']))
            remarks = sanitize(get_val(['Remark', 'remark', '비고']))
            is_tax_free = (str(get_val(['IsTaxFree', '면세여부'])) in ['TRUE', '면세'])

            # Determine Target ID
            target_id = None
            if model_no and model_no in model_no_map:
                target_id = model_no_map[model_no]
            elif model_name and normalize_name(model_name) in name_map:
                ids = name_map[normalize_name(model_name)]
                if ids: target_id = ids[0]

            if target_id:
                # Update
                con.run(
                    """UPDATE products 
                       SET model_name = :m_name,
                           model_no = :m_no,
                           category_id = COALESCE(:cat_id, category_id),
                           brand = :brand,
                           b2b_price = :b2b, 
                           consumer_price = :consumer, 
                           supply_price = :supply,
                           quantity_per_carton = :qty,
                           description = COALESCE(:desc, description),
                           detail_url = COALESCE(:detail, detail_url),
                           image_url = COALESCE(:img, image_url),
                           manufacturer = :manuf,
                           origin = :origin,
                           product_spec = :spec,
                           product_options = :opts,
                           remarks = :rem,
                           is_tax_free = :tax,
                           shipping_fee = :ship,
                           shipping_fee_individual = :ship_ind,
                           shipping_fee_carton = :ship_carton,
                           stock_quantity = :stock,
                           updated_at = CURRENT_TIMESTAMP
                       WHERE id = :id""",
                    m_name=model_name, m_no=model_no, cat_id=cat_id, brand=brand,
                    b2b=b2b_price, consumer=consumer_price, supply=supply_price, 
                    qty=quantity_per_carton, desc=description, detail=detail_url, 
                    img=image_url, id=target_id,
                    manuf=manufacturer, origin=origin, spec=product_spec, opts=product_options,
                    rem=remarks, tax=is_tax_free, ship=shipping_fee, ship_ind=shipping_fee_ind,
                    ship_carton=shipping_fee_carton, stock=stock_quantity
                )
                print(f"Updated Product {target_id} ({model_name})")
                updated_count += 1
            else:
                # Insert
                con.run(
                    """INSERT INTO products (
                        category_id, brand, model_name, description, image_url, b2b_price, stock_quantity,
                        detail_url, consumer_price, supply_price, quantity_per_carton, shipping_fee,
                        manufacturer, origin, is_tax_free, shipping_fee_individual, shipping_fee_carton,
                        product_options, model_no, product_spec, remarks, is_available
                    ) VALUES (
                        :cat_id, :brand, :model_name, :desc, :img, :b2b, :stock,
                        :detail, :consumer, :supply, :qty, :ship,
                        :manuf, :origin, :tax, :ship_ind, :ship_carton,
                        :opts, :m_no, :spec, :rem, true
                    )""",
                    cat_id=cat_id, brand=brand, model_name=model_name, desc=description, img=image_url,
                    b2b=b2b_price, stock=stock_quantity, detail=detail_url, consumer=consumer_price,
                    supply=supply_price, qty=quantity_per_carton, ship=shipping_fee, manuf=manufacturer,
                    origin=origin, tax=is_tax_free, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    opts=product_options, m_no=model_no, spec=product_spec, rem=remarks
                )
                print(f"Inserted New Product ({model_name})")
                inserted_count += 1

        except Exception as e:
            print(f"Error processing row for {model_name}: {e}")
            error_count += 1

    con.close()
    print(f"Range Processing Complete. Inserted: {inserted_count}, Updated: {updated_count}, Errors: {error_count}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Register products from Excel range')
    parser.add_argument('--start', type=int, required=True, help='Start Row')
    parser.add_argument('--end', type=int, required=True, help='End Row')
    
    args = parser.parse_args()
    register_range(args.start, args.end)
