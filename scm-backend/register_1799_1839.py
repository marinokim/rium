import sys
import os
import ssl
sys.path.append(os.path.join(os.getcwd(), 'pylib'))

import openpyxl
import pg8000.native

# Helper to sanitize and parse
def sanitize(val, allow_html=False):
    if val is None:
        return None
    s = str(val).replace('_x000D_', '').replace('\r', '').strip()
    if not allow_html:
        s = s.replace('"', '')
        if s.startswith('<') or 'img src' in s:
            return None
    return s

def parse_price(val):
    if val is None:
        return 0
    s = str(val).replace(',', '').replace('원', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

def normalize_name(name):
    if not name:
        return ""
    return str(name).replace('_x000D_', '').replace('\r', '').replace('\n', '').strip().lower()

def register_1799_1839():
    # DB Connection
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE

    print("Connecting to DB...")
    con = pg8000.native.Connection(
        user="postgres.zgqnhbaztgpekerhxwbc",
        password="gmlrhks0528&*",
        host="aws-1-ap-northeast-2.pooler.supabase.com",
        port=5432,
        database="postgres",
        ssl_context=ssl_context
    )

    print("Pre-fetching existing products...")
    all_products = con.run("SELECT id, model_no, model_name FROM products")
    
    name_map = {}
    model_no_map = {}
    
    for pid, mno, mname in all_products:
        if mno:
            clean_mno = str(mno).strip()
            model_no_map[clean_mno] = pid
        if mname:
            n_name = normalize_name(mname)
            if n_name not in name_map:
                name_map[n_name] = []
            name_map[n_name].append(pid)
    
    for n in name_map:
        name_map[n].sort()

    print(f"Loaded {len(all_products)} products into map.")

    excel_path = './excel/aron_product_upload_consolidated.xlsx'
    print(f"Loading Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active

    # Get headers
    headers = {}
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    for idx, col_name in enumerate(header_row):
        if col_name:
            headers[str(col_name).strip()] = idx

    print(f"Headers found: {list(headers.keys())}")

    min_r = 1799
    max_r = 1839
    print(f"Processing rows {min_r} to {max_r}...")
    
    row_idx = min_r - 1 
    updated_count = 0
    
    for row in sheet.iter_rows(min_row=min_r, max_row=max_r, values_only=True):
        row_idx += 1
        
        def get_val(possible_headers):
            for h in possible_headers:
                if h in headers:
                    return row[headers[h]]
            return None

        def truncate(val, length=255):
            if val and len(val) > length:
                return val[:length]
            return val

        brand = sanitize(get_val(['Brand', '브랜드']))
        model_name_raw = get_val(['ModelName', '모델명', '상품명'])
        model_name = truncate(sanitize(model_name_raw))
        model_no = truncate(sanitize(get_val(['ModelNo', '모델번호'])))
        
        # Prices
        b2b_price = parse_price(get_val(['B2BPrice', '실판매가', '판매가', 'B2B가']))
        consumer_price = parse_price(get_val(['ConsumerPrice', '소비자가', '소가']))
        supply_price = parse_price(get_val(['SupplyPrice', '공급가', '매입가']))
        if supply_price == 0 and b2b_price > 0:
            supply_price = b2b_price
            
        quantity_per_carton = parse_price(get_val(['QuantityPerCarton', '카톤수량']))
        if quantity_per_carton == 0: quantity_per_carton = 1
            
        description = sanitize(get_val(['Description', '상세설명']))
        if description: description = description.replace('_x000D_', '\n')
             
        # Allow HTML for DetailURL
        detail_url = sanitize(get_val(['DetailURL', '상세페이지URL']), allow_html=True) 
        
        image_url = truncate(sanitize(get_val(['ImageURL', '이미지URL'])))

        stock_quantity = parse_price(get_val(['Stock', '재고']))
        shipping_fee = parse_price(get_val(['ShippingFee', '배송비']))
        shipping_fee_ind = parse_price(get_val(['ShippingFeeIndividual', '개별배송비']))
        if shipping_fee_ind == 0 and shipping_fee > 0: shipping_fee_ind = shipping_fee
        shipping_fee_carton = parse_price(get_val(['ShippingFeeCarton', '카톤배송비']))
        manufacturer = sanitize(get_val(['Manufacturer', '제조사']))
        origin = sanitize(get_val(['Origin', '원산지']))
        product_spec = sanitize(get_val(['ProductSpec', '제품규격']))
        product_options = sanitize(get_val(['ProductOptions', '옵션']))
        remarks = sanitize(get_val(['Remark', 'remark', '비고']))
        is_tax_free = (str(get_val(['IsTaxFree', '면세여부'])) in ['TRUE', '면세'])

        category_name = sanitize(get_val(['Category', '카테고리']))
        
        # Category Logic
        cat_id = 10 # Default to Living (KAUF items mostly)
        
        cat_str = str(category_name).lower() if category_name else ""
        
        # Priority mapping
        if 'coffee' in cat_str:
            cat_id = 5 # Food
        elif 'mobile' in cat_str or '보조배터리' in str(model_name) or 'power bank' in str(model_name).lower():
            cat_id = 2 # Mobile (Tech)
        elif 'kitchen' in cat_str:
            cat_id = 11
        elif 'gift' in cat_str:
            cat_id = 8
        elif 'beauty' in cat_str:
            cat_id = 3

        if not model_name:
            print(f"Row {row_idx}: Skipping (No Model Name)")
            continue

        # Check existing
        target_id = None
        if model_no and model_no in model_no_map:
            target_id = model_no_map[model_no]
        elif model_name and normalize_name(model_name) in name_map:
            ids = name_map[normalize_name(model_name)]
            if ids: target_id = ids[0]

        try:
            if target_id:
                # Update
                con.run(
                    """UPDATE products 
                       SET model_name = :m_name,
                           model_no = :m_no,
                           category_id = :cat_id,
                           b2b_price = :b2b, 
                           consumer_price = :consumer, 
                           supply_price = :supply,
                           quantity_per_carton = :qty,
                           description = COALESCE(:desc, description),
                           detail_url = COALESCE(:detail, detail_url),
                           image_url = COALESCE(:img, image_url),
                           shipping_fee = :ship,
                           shipping_fee_individual = :ship_ind,
                           shipping_fee_carton = :ship_carton,
                           brand = :brand,
                           updated_at = CURRENT_TIMESTAMP
                       WHERE id = :id""",
                    m_name=model_name, m_no=model_no, cat_id=cat_id,
                    b2b=b2b_price, consumer=consumer_price, supply=supply_price, 
                    qty=quantity_per_carton, desc=description, detail=detail_url, 
                    img=image_url, id=target_id,
                    ship=shipping_fee, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    brand=brand
                )
                print(f"Row {row_idx}: Updated Product {target_id} ({model_name}) - Cat: {cat_id}")
            else:
                # Insert
                con.run(
                    """INSERT INTO products (
                        category_id, brand, model_name, description, image_url, b2b_price, stock_quantity,
                        detail_url, consumer_price, supply_price, quantity_per_carton, shipping_fee,
                        manufacturer, origin, is_tax_free, shipping_fee_individual, shipping_fee_carton,
                        product_options, model_no, product_spec, remarks, is_available
                    ) VALUES (
                        :cat_id, :brand, :model_name, :desc, :img, :b2b, :stock,
                        :detail, :consumer, :supply, :qty, :ship,
                        :manuf, :origin, :tax, :ship_ind, :ship_carton,
                        :opts, :m_no, :spec, :rem, true
                    )""",
                    cat_id=cat_id, brand=brand, model_name=model_name, desc=description, img=image_url,
                    b2b=b2b_price, stock=stock_quantity, detail=detail_url, consumer=consumer_price,
                    supply=supply_price, qty=quantity_per_carton, ship=shipping_fee, manuf=manufacturer,
                    origin=origin, tax=is_tax_free, ship_ind=shipping_fee_ind, ship_carton=shipping_fee_carton,
                    opts=product_options, m_no=model_no, spec=product_spec, rem=remarks
                )
                print(f"Row {row_idx}: Inserted New Product ({model_name}) - Cat: {cat_id}")
            
            updated_count += 1

        except Exception as e:
            print(f"Row {row_idx}: Error - {e}")

    print(f"Done. Processed {updated_count} items.")
    con.close()

if __name__ == "__main__":
    register_1799_1839()
